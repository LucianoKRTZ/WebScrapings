import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pandas as pd
from weasyprint import HTML
from pathlib import Path
from datetime import datetime
import os

pathAcessos = "C:\\__Programas__\\__config__\\acessos.xlsx"
pathOutput = "C:\\Macros\\Acoes"

data = datetime.today()

class DaoUtils:
    def obterAcessos(self, plataforma):
        excel = pd.read_excel(pathAcessos, sheet_name='acessos')
        for linha in excel.itertuples():
            if linha[1] == plataforma:
                usuario = linha[2]
                senha = linha[3]

        return usuario, senha
    
    def obterDestinatarioEmail(self, programa):
        excel = pd.read_excel(pathAcessos, sheet_name='destinatarios')
        destinatarios = []
        for linha in excel.itertuples():
            if linha[1] == programa:
                destinatarios.append(linha[3])
        return destinatarios

    def organizarPastaMacroAcoes(self):
        print(f"\n\n\n=> Organizando arquivos na pasta: {pathOutput}")

        for arq in os.listdir(pathOutput):
            pathArq = os.path.join(pathOutput, arq)
            if os.path.isfile(pathArq):
                print(f"    ->Organizando arquivo: {arq}")

                dataArq = arq.split('-')[-1].replace('.'+arq.split('.')[-1],'')
                dataArq = datetime.strptime(dataArq, '%d.%m.%Y')

                if dataArq.__format__('%d%m%Y') != data.__format__('%d%m%Y'):
                    pathDst = os.path.join(pathOutput, dataArq.strftime('%Y'), dataArq.strftime('%m'))
                    if not os.path.exists(pathDst):
                        os.makedirs(pathDst)
                    novoCaminho = os.path.join(pathDst, arq)
                    os.rename(pathArq, novoCaminho)

    def enviarEmail(self, destinatarios, assunto, mensagem):

        usuario, senha = self.obterAcessos("emailMacros")
        print(usuario, senha)

        smtp_server = "smtp.office365.com"
        smtp_port = 587

        msgEmail = f"""
Bom dia!

{mensagem}

Atenciosamente,
Robô do Luciano.
"""


        msg = MIMEMultipart()
        msg['From'] = usuario
        msg['To'] = destinatarios
        msg['Subject'] = assunto
        msg.attach(MIMEText(mensagem, 'plain'))

        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(usuario, senha)
            server.sendmail(usuario, destinatarios, msg.as_string())
            server.quit()
            print("E-mail enviado com sucesso!")
        except Exception as e:
            print(f"Falha ao enviar e-mail: {e}")

    def converterHTML2PDF(self, conteudoHTML, diretorioSaida, nomeArquivo):

        caminhoSaida = Path(diretorioSaida) / f"{nomeArquivo}.pdf"
        HTML(string=conteudoHTML).write_pdf(caminhoSaida)
        print(f"PDF gerado em: {caminhoSaida}")
        return caminhoSaida
    
    def adicionarMenuCompras(self):

        for arq in os.listdir(pathOutput):
            pathArq = os.path.join(pathOutput, arq)
            if os.path.isfile(pathArq) and arq.endswith('.xlsx') and \
            data.__format__("%d.%m.%Y") in arq:
                print(f"    ->Editando arquivo: {arq}")
                # Abrir a planilha
                xls = pd.ExcelFile(pathArq)
                abas = xls.sheet_names
                abas_origem = []
                if 'Ações' in abas:
                    abas_origem.append('Ações')
                if 'FII-Fiagro' in abas:
                    abas_origem.append('FII-Fiagro')
                df_list = []
                for aba in abas_origem:
                    df = pd.read_excel(pathArq, sheet_name=aba)
                    colunas_necessarias = ['Ticker', 'Segmento', 'Valor Atual', 'Média de Dividendos (24m)']
                    df = df[[col for col in colunas_necessarias if col in df.columns]]
                    df_list.append(df)
                if df_list:
                    import openpyxl
                    df_menu = pd.concat(df_list, ignore_index=True)
                    df_menu['Quantidade'] = 0
                    # Adiciona as colunas como fórmulas do Excel
                    df_menu['Valor Total'] = None  # Placeholder for formula
                    df_menu['Estimativa Dividendos'] = None  # Placeholder for formula
                    colunas_finais = ['Ticker', 'Segmento', 'Valor Atual', 'Média de Dividendos (24m)', 'Quantidade', 'Valor Total', 'Estimativa Dividendos']
                    df_menu = df_menu[[col for col in colunas_finais if col in df_menu.columns]]
                    with pd.ExcelWriter(pathArq, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_menu.to_excel(writer, sheet_name='Menu-Compras', index=False)
                    # Aplicar fórmulas nas colunas 'Valor Total' e 'Estimativa Dividendos'
                    wb = openpyxl.load_workbook(pathArq)
                    ws = wb['Menu-Compras']
                    for row in range(2, ws.max_row + 1):
                        ws[f'F{row}'] = f'=C{row}*E{row}'  # Valor Total
                        ws[f'G{row}'] = f'=D{row}*E{row}'  # Estimativa Dividendos
                    # Adiciona resumo ao final da tabela
                    resumo_row = ws.max_row + 2
                    
                    ws[f'B{resumo_row}'] = 'Valor total'
                    ws[f'C{resumo_row}'] = f'=SUM(F2:F{ws.max_row-2})'
                    
                    ws[f'B{resumo_row+1}'] = 'Total de dividendos estimado'
                    ws[f'C{resumo_row+1}'] = f'=SUM(G2:G{ws.max_row-2})'
                    
                    # Aplicar formatação monetária na aba Menu-Compras
                    self.formatarAbaMenuCompras(wb, 'Menu-Compras')
                    
                    # Ajustar largura das colunas da aba Menu-Compras
                    self.ajustarLargurasAba(wb, 'Menu-Compras')
                    wb.save(pathArq)
                    
                    print(f"    ->Aba 'Menu-Compras' criada/atualizada em: {arq}")
                print(f"    ->Arquivo editado: {arq}")

    def editarPlanilhaAcoes(self, pathExcel):
        print(f"    -> Editando arquivo: {pathExcel}")

        self.formatarPlanilhaAcoes(pathExcel)
        self.alinharCelulasNumericas(pathExcel)
        self.ajustarLarguraColunas(pathExcel)

    def formatarPlanilhaAcoes(self, pathExcel):
        """
        Formata valores monetários nas colunas especificadas
        """
        import openpyxl
        from openpyxl import styles
        
        print(f"    -> Formatando valores monetários em: {pathExcel}")
        
        # Colunas que devem receber formatação monetária
        colunas_monetarias = ['Valor Atual', 'Valor min.', 'Valor max.', 'Média de Dividendos (24m)']
        
        wb = openpyxl.load_workbook(pathExcel)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"        -> Formatando valores da aba: {sheet_name}")
            
            # Encontrar os índices das colunas monetárias
            colunas_indices = {}
            if ws.max_row > 0:
                # Verificar a primeira linha (cabeçalhos)
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value and str(cell_value) in colunas_monetarias:
                        colunas_indices[col] = str(cell_value)
            
            # Formatar as células das colunas monetárias
            for col_idx, col_name in colunas_indices.items():
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            # Converter string monetária para número
                            valor_numerico = self.extrairValorNumerico(str(cell.value))
                            if valor_numerico is not None:
                                cell.value = valor_numerico
                                cell.number_format = 'R$ #,##0.00'
                                cell.alignment = styles.Alignment(horizontal='right', vertical='center')
        
        wb.save(pathExcel)
        print(f"    -> Formatação monetária aplicada com sucesso!")
    
    def alinharCelulasNumericas(self, pathExcel):
        """
        Alinha todas as células com números à direita, exceto aquelas com letras ou datas (com /)
        """
        import openpyxl
        from openpyxl import styles
        import re
        
        print(f"    -> Alinhando células numéricas em: {pathExcel}")
        
        wb = openpyxl.load_workbook(pathExcel)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"        -> Alinhando células da aba: {sheet_name}")
            
            # Percorrer todas as células
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_value_str = str(cell.value)
                        
                        # Verificar se é um número (int, float ou string numérica)
                        if self.ehCelulaNumerica(cell.value, cell_value_str):
                            # Aplicar alinhamento à direita apenas se não for data ou texto
                            if not self.ehData(cell_value_str) and not self.contemLetras(cell_value_str):
                                cell.alignment = styles.Alignment(horizontal='right', vertical='center')
        
        wb.save(pathExcel)
        print(f"    -> Alinhamento de células numéricas aplicado com sucesso!")
    
    def ehCelulaNumerica(self, value, value_str):
        """
        Verifica se o valor é numérico (int, float ou string que representa número)
        """
        import re

        # Se já é int ou float
        if isinstance(value, (int, float)):
            return True
        
        # Se é string, verificar se representa um número
        if isinstance(value, str):
            # Remover espaços
            value_clean = value_str.strip()
            
            # Verificar se é uma fórmula (começar com =)
            if value_clean.startswith('='):
                return True
            
            # Verificar se é um percentual (contém %)
            if '%' in value_clean:
                # Remove % e verifica se o resto é numérico
                valor_sem_percent = value_clean.replace('%', '').replace(',', '.').replace('-', '').strip()
                try:
                    float(valor_sem_percent)
                    return True
                except ValueError:
                    pass
            
            # Verificar se é um número (com ou sem símbolos monetários)
            # Remove R$, espaços, vírgulas e pontos para testar
            test_value = re.sub(r'[R$\s,.]', '', value_clean)
            if test_value.replace('-', '').isdigit():
                return True
                
            # Verificar se é número decimal
            return self.ehNumeroDecimal(value_clean)
        
        return False
    
    def ehNumeroDecimal(self, texto):
        """
        Verifica se o texto representa um número decimal
        """
        import re
        # Padrão para números decimais (com vírgula ou ponto, podendo ser negativo, com % ou R$)
        padrao_decimal = r'^[R$\s]*[-]?[\d]+[,.]?[\d]*[%]?$'
        return bool(re.match(padrao_decimal, texto.strip()))
    
    def ehData(self, texto):
        """
        Verifica se o texto contém barra (/), indicando que pode ser uma data
        """
        return '/' in texto
    
    def contemLetras(self, texto):
        """
        Verifica se o texto contém letras (exceto R$ e % que são símbolos)
        """
        import re
        # Remove símbolos monetários, percentuais e números, verifica se sobram letras
        texto_sem_simbolos = re.sub(r'[R$%\s\d,.-]', '', texto)
        return bool(re.search(r'[a-zA-ZÀ-ÿ]', texto_sem_simbolos))

    def alinharCelulasAba(self, workbook, sheet_name):
        """
        Alinha células numéricas de uma aba específica à direita
        """
        from openpyxl import styles
        
        ws = workbook[sheet_name]
        
        # Percorrer todas as células da aba
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_value_str = str(cell.value)
                    
                    # Verificar se é um número (int, float ou string numérica)
                    if self.ehCelulaNumerica(cell.value, cell_value_str):
                        # Aplicar alinhamento à direita apenas se não for data ou texto
                        if not self.ehData(cell_value_str) and not self.contemLetras(cell_value_str):
                            cell.alignment = styles.Alignment(horizontal='right', vertical='center')
    
    def extrairValorNumerico(self, texto):
        """
        Extrai o valor numérico de uma string monetária
        Exemplo: "R$0,07938354" -> 0.08 (arredondado para 2 casas decimais)
        """
        import re
        
        # Remover símbolos de moeda e espaços
        texto_limpo = re.sub(r'[R$\s]', '', texto)
        
        # Substituir vírgula por ponto para conversão decimal
        texto_limpo = texto_limpo.replace(',', '.')
        
        try:
            valor = float(texto_limpo)
            # Arredondar para 2 casas decimais
            return round(valor, 2)
        except ValueError:
            return None

    def formatarAbaMenuCompras(self, workbook, sheet_name):
        """
        Aplica formatação monetária específica para a aba Menu-Compras
        """
        from openpyxl import styles
        
        ws = workbook[sheet_name]
        
        # Colunas que devem receber formatação monetária no Menu-Compras
        colunas_monetarias = ['Valor Atual', 'Média de Dividendos (24m)', 'Valor Total', 'Estimativa Dividendos']
        
        # Encontrar os índices das colunas monetárias
        colunas_indices = {}
        if ws.max_row > 0:
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value) in colunas_monetarias:
                    colunas_indices[col] = str(cell_value)
        
        # Formatar as células das colunas monetárias
        for col_idx, col_name in colunas_indices.items():
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        # Se for uma célula com fórmula, aplicar apenas a formatação
                        if str(cell.value).startswith('='):
                            cell.number_format = 'R$ #,##0.00'
                            cell.alignment = styles.Alignment(horizontal='right', vertical='center')
                        else:
                            # Se for um valor, converter e formatar
                            valor_numerico = self.extrairValorNumerico(str(cell.value))
                            if valor_numerico is not None:
                                cell.value = valor_numerico
                                cell.number_format = 'R$ #,##0.00'
                                cell.alignment = styles.Alignment(horizontal='right', vertical='center')
        
        # Formatar as células de resumo também
        for row in range(ws.max_row - 1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and str(cell.value).startswith('='):
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = styles.Alignment(horizontal='right', vertical='center')

    def ajustarLarguraColunas(self, pathExcel):
        """
        Ajusta automaticamente a largura das colunas em todas as abas da planilha
        para que fiquem com a largura do maior texto presente na coluna
        """
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        print(f"    -> Ajustando largura das colunas em: {pathExcel}")
        
        wb = openpyxl.load_workbook(pathExcel)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"        -> Ajustando colunas da aba: {sheet_name}")
            
            # Dicionário para armazenar a largura máxima de cada coluna
            column_widths = {}
            
            # Percorrer todas as células para encontrar o maior conteúdo de cada coluna
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Converter o valor para string e calcular o comprimento
                        cell_length = len(str(cell.value))
                        column_letter = get_column_letter(cell.column)
                        
                        # Atualizar a largura máxima da coluna se necessário
                        if column_letter not in column_widths or cell_length > column_widths[column_letter]:
                            column_widths[column_letter] = cell_length
            
            # Aplicar as larguras calculadas às colunas
            for column_letter, width in column_widths.items():
                # Adicionar uma margem extra (mínimo 10, máximo 50)
                adjusted_width = min(max(width + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(pathExcel)
        print(f"    -> Larguras das colunas ajustadas com sucesso!")

    def ajustarLargurasAba(self, workbook, sheet_name):
        """
        Ajusta automaticamente a largura das colunas de uma aba específica
        """
        from openpyxl.utils import get_column_letter
        
        ws = workbook[sheet_name]
        
        # Dicionário para armazenar a largura máxima de cada coluna
        column_widths = {}
        
        # Percorrer todas as células para encontrar o maior conteúdo de cada coluna
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Converter o valor para string e calcular o comprimento
                    cell_length = len(str(cell.value))
                    column_letter = get_column_letter(cell.column)
                    
                    # Atualizar a largura máxima da coluna se necessário
                    if column_letter not in column_widths or cell_length > column_widths[column_letter]:
                        column_widths[column_letter] = cell_length
        
        # Aplicar as larguras calculadas às colunas
        for column_letter, width in column_widths.items():
            # Adicionar uma margem extra (mínimo 10, máximo 50)
            adjusted_width = min(max(width + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width




